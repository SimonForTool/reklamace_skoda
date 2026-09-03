"""
Datový model a úložiště pro Reklamace — správa reklamací pro více značek
(ŠKODA, PORSCHE), každá se samostatnou číselnou řadou.

Reklamace prochází 4 nezávislými fázemi (WETSy, GETAC, ACTIA IMEA,
Ostatní dodavatel), z nichž každá se eviduje samostatně (stav, datum,
poznámka, odpovědná osoba, přílohy). Interní číslo má formát
<prefix značky>-<rok>-<pořadové číslo v daném roce>, rok se odvozuje
z data přijetí. Číselná řada se počítá zvlášť pro každou značku.
"""
from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from pathlib import Path
from werkzeug.utils import secure_filename

DATA_DIR = Path("data")
DATA_PATH = DATA_DIR / "reklamace.json"
UPLOAD_DIR = Path("uploads") / "reklamace"

BRANDS = {
    "skoda":   {"label": "ŠKODA",   "prefix": "SKO", "logo": "skoda.webp"},
    "porsche": {"label": "PORSCHE", "prefix": "POR", "logo": "porsche.png"},
}
DEFAULT_BRAND = "skoda"

FAZE_DEFS = [
    {"key": "wetsy",   "label": "WETSy"},
    {"key": "getac",   "label": "GETAC"},
    {"key": "actia",   "label": "ACTIA IMEA"},
    {"key": "ostatni", "label": "Ostatní dodavatel"},
]
FAZE_KEYS = [f["key"] for f in FAZE_DEFS]
FAZE_LABEL_BY_KEY = {f["key"]: f["label"] for f in FAZE_DEFS}

STAVY = ["nezahájeno", "probíhá", "hotovo"]

# Katalog kroků procesu — společný pro všechny 4 záložky (WETSy, GETAC,
# ACTIA IMEA, Ostatní dodavatel). V appce se řádky přidávají přes rolovací
# seznam; každá volba předvyplní obvyklou dobu v minutách (editovatelnou).
OSTATNI_KATALOG_LABEL = "Ostatní - doplnit"
KROK_KATALOG = [
    {"nazev": "Telefonický příjem reklamace",           "minuty": 15},
    {"nazev": "Odchozí informační e-mail",               "minuty": 15},
    {"nazev": "Příchozí e-mail, založení složky",        "minuty": 30},
    {"nazev": "Založení ticket",                         "minuty": 60},
    {"nazev": "Komunikace VAT & ŠPC (č. fa z VW)",       "minuty": 30},
    {"nazev": "Komunikace se zák. (e-mail, tel.)",       "minuty": 30},
    {"nazev": "Odeslání reklamace",                      "minuty": 60},
    {"nazev": "Správa ticketu",                          "minuty": 30},
    {"nazev": "Potvrzení o ukončení reklamace",          "minuty": 15},
    {"nazev": "Komunikace s dodavatelem",                "minuty": 30},
    {"nazev": "Příjem reklamace",                        "minuty": 30},
    {"nazev": "Uzavření ticketu",                        "minuty": 25},
    {"nazev": OSTATNI_KATALOG_LABEL,                     "minuty": 15},
]
KROK_KATALOG_BY_NAZEV = {k["nazev"]: k for k in KROK_KATALOG}
EXTRA_KROK_MINUTY_DEFAULT = 15

STAV_LABELS_HISTORIE = {
    "nová": "Otevřeno", "probíhá": "Probíhá", "vyřízeno": "Uzavřeno", "fakturováno": "Fakturováno",
}

# Export VAT do Excelu (podklad pro fakturaci) — jeden řádek na uzavřenou
# reklamaci, souhrnné hodiny za každého dodavatele (bez rozpisu kroků).
EXPORT_FAZE_SLOUPCE = [
    ("wetsy",   "WETSy (hod.)"),
    ("getac",   "GETAC (hod.)"),
    ("actia",   "ACTIA IME (hod.)"),
    ("ostatni", "Ostatní dodavatel (hod.)"),
]

NASTAVENI_DEFAULTS = {
    "skoda":   {"prevod_hodin": 431},
    "porsche": {"prevod_hodin": 0},
}

DOPRAVCI = ["Česká pošta", "DHL", "Geis", "PPL", "GLS", "Zásilkovna", "DPD", "Toptrans", "Jiný"]


def get_nastaveni(brand: str) -> dict:
    data = load_all()
    ulozene = data.get("nastaveni", {}).get(brand, {})
    default = NASTAVENI_DEFAULTS.get(brand, {"prevod_hodin": 0})
    return {**default, **ulozene}


def update_nastaveni(brand: str, patch: dict) -> dict:
    data = load_all()
    nastaveni = data.setdefault("nastaveni", {}).setdefault(brand, dict(get_nastaveni(brand)))
    if "prevod_hodin" in patch:
        try:
            nastaveni["prevod_hodin"] = float(patch["prevod_hodin"])
        except (TypeError, ValueError):
            pass
    save_all(data)
    return nastaveni


def potvrdit_fakturaci(brand: str, cisla: list | None = None) -> dict:
    """Označí uzavřené (a dosud nefakturované) reklamace dané značky jako
    Fakturováno a odečte jejich odpracované hodiny z rozpočtu PŘEVOD hodin
    — zůstatek se stává počátkem pro příště. Pokud je zadán výběr cisla,
    zahrnou se jen ty (a stále jen pokud jsou skutečně uzavřené); jinak
    všechny aktuálně uzavřené."""
    data = load_all()
    items = [i for i in data.get("items", {}).values()
             if i.get("znacka", DEFAULT_BRAND) == brand and overall_status(i) == "vyřízeno"]
    if cisla is not None:
        vybrana = set(cisla)
        items = [i for i in items if i["cislo"] in vybrana]
    total_minutes = sum(reklamace_total_minutes(i) for i in items)
    total_hodin = round(total_minutes / 60, 2)

    for i in items:
        i["fakturovano"] = True

    nastaveni = data.setdefault("nastaveni", {}).setdefault(brand, dict(get_nastaveni(brand)))
    novy_zustatek = round(nastaveni.get("prevod_hodin", 0) - total_hodin, 2)
    nastaveni["prevod_hodin"] = novy_zustatek
    save_all(data)

    return {
        "pocet_reklamaci": len(items),
        "celkem_hodin": total_hodin,
        "novy_zustatek": novy_zustatek,
    }


def _empty_kroky() -> list:
    return []


def _empty_faze(key: str) -> dict:
    f = {
        "stav": "nezahájeno",
        "datum": None,
        "poznamka": "",
        "odpovedna_osoba": "",
        "prilohy": [],
        "kroky": _empty_kroky(),
        # Náklady a fakturace
        "naklady_fortool": None,
        "dopravce": "",
        "vydana_faktura": None,
        "uhrada_potvrzena": False,
        "datum_uhrady": None,
        "poznamky_faktura": "",
    }
    if key == "ostatni":
        f["dodavatel"] = ""
    return f


def faze_total_minutes(faze: dict) -> float:
    return sum((k.get("minuty") or 0) for k in faze.get("kroky", []))


def reklamace_total_minutes(item: dict) -> float:
    return sum(faze_total_minutes(item["faze"][k]) for k in FAZE_KEYS)


def reklamace_naklady_celkem(item: dict) -> float:
    return sum((item["faze"][k].get("naklady_fortool") or 0) for k in FAZE_KEYS)


def reklamace_faktura_celkem(item: dict) -> float:
    return sum((item["faze"][k].get("vydana_faktura") or 0) for k in FAZE_KEYS)


def reklamace_faktura_neuhrazeno(item: dict) -> float:
    return sum(
        (item["faze"][k].get("vydana_faktura") or 0)
        for k in FAZE_KEYS
        if (item["faze"][k].get("vydana_faktura") or 0) > 0 and not item["faze"][k].get("uhrada_potvrzena")
    )


def reklamace_uhrada_stav(item: dict) -> str:
    faktury = [item["faze"][k] for k in FAZE_KEYS if (item["faze"][k].get("vydana_faktura") or 0) > 0]
    if not faktury:
        return "—"
    if all(f.get("uhrada_potvrzena") for f in faktury):
        return "Uhrazeno"
    if any(f.get("uhrada_potvrzena") for f in faktury):
        return "Částečně uhrazeno"
    return "Čeká na úhradu"


def reklamace_poznamky_souhrn(item: dict) -> str:
    casti = []
    for k in FAZE_KEYS:
        text = (item["faze"][k].get("poznamky_faktura") or "").strip()
        if text:
            casti.append(f"{FAZE_LABEL_BY_KEY[k]}: {text}")
    return " | ".join(casti)


def minutes_to_hours(minutes: float) -> float:
    return round(minutes / 60, 1)


def business_days_between(start: date, end: date) -> int:
    """Počet pracovních dnů (po-pá) mezi start (včetně) a end (bez), min. 0."""
    if not start or not end or end <= start:
        return 0
    days = 0
    d = start
    while d < end:
        if d.weekday() < 5:
            days += 1
        d += timedelta(days=1)
    return days


def _sync_uzavreno(item: dict):
    """Datum ukončení je editovatelné ručně na detailu reklamace — jen
    doplní výchozí hodnotu při uzavření všech fází, nikdy ho nemaže."""
    if overall_status(item) == "vyřízeno" and not item.get("uzavreno"):
        item["uzavreno"] = date.today().isoformat()


def aktualni_faze_label(item: dict) -> str:
    probihajici = [FAZE_LABEL_BY_KEY[k] for k in FAZE_KEYS if item["faze"][k]["stav"] == "probíhá"]
    if probihajici:
        return ", ".join(probihajici)
    if overall_status(item) == "vyřízeno":
        return "—"
    return "Nezahájeno"


def _empty_store() -> dict:
    return {"counters": {}, "items": {}}


def load_all() -> dict:
    if not DATA_PATH.exists():
        return _empty_store()
    return json.loads(DATA_PATH.read_text())


def save_all(data: dict):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    DATA_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2))


def list_all(brand: str | None = None) -> list:
    data = load_all()
    items = list(data.get("items", {}).values())
    if brand:
        items = [i for i in items if i.get("znacka", DEFAULT_BRAND) == brand]
    items.sort(key=lambda i: i["vytvoreno"], reverse=True)
    for item in items:
        item["celkovy_stav"] = overall_status(item)
        item["celkem_hodin"] = minutes_to_hours(reklamace_total_minutes(item))
    return items


def overall_status(item: dict) -> str:
    # Fakturováno = uzavřená reklamace už zahrnutá do potvrzené fakturace
    # (viz potvrdit_fakturaci) — trvalý koncový stav, dokud se ručně
    # nezruší. Ruční datum ukončení reklamace (na detailu) rozhoduje
    # přednostně před fázemi — ne každá reklamace prochází všemi 4
    # fázemi, takže čekat na "hotovo" u všech čtyř by u řady reklamací
    # nikdy nenastalo.
    if item.get("fakturovano"):
        return "fakturováno"
    if item.get("uzavreno"):
        return "vyřízeno"
    stavy = [item["faze"][k]["stav"] for k in FAZE_KEYS]
    if all(s == "hotovo" for s in stavy):
        return "vyřízeno"
    if any(s != "nezahájeno" for s in stavy):
        return "probíhá"
    return "nová"


def get_reklamace(cislo: str) -> dict | None:
    data = load_all()
    item = data.get("items", {}).get(cislo)
    if item:
        item["celkovy_stav"] = overall_status(item)
        item["celkem_hodin"] = minutes_to_hours(reklamace_total_minutes(item))
    return item


def new_reklamace(brand: str, servisni_partner: str, kontaktni_osoba: str, datum_prijeti: str) -> dict:
    if brand not in BRANDS:
        raise ValueError(f"Neznámá značka: {brand}")
    data = load_all()
    try:
        year = date.fromisoformat(datum_prijeti).year
    except (TypeError, ValueError):
        year = date.today().year
        datum_prijeti = date.today().isoformat()

    counters = data.setdefault("counters", {})
    counter_key = f"{brand}:{year}"
    counters[counter_key] = counters.get(counter_key, 0) + 1
    prefix = BRANDS[brand]["prefix"]
    cislo = f"{prefix}-{year}-{counters[counter_key]:03d}"

    item = {
        "cislo": cislo,
        "znacka": brand,
        "servisni_partner": servisni_partner.strip(),
        "kontaktni_osoba": kontaktni_osoba.strip(),
        "datum_prijeti": datum_prijeti,
        "uzavreno": None,
        "fakturovano": False,
        "odpovedna_osoba": "",
        "vytvoreno": datetime.now().isoformat(timespec="seconds"),
        "faze": {k: _empty_faze(k) for k in FAZE_KEYS},
    }
    data.setdefault("items", {})[cislo] = item
    save_all(data)
    item["celkovy_stav"] = overall_status(item)
    item["celkem_hodin"] = minutes_to_hours(reklamace_total_minutes(item))
    return item


def update_header(cislo: str, patch: dict) -> dict | None:
    data = load_all()
    item = data.get("items", {}).get(cislo)
    if not item:
        return None
    for key in ("servisni_partner", "kontaktni_osoba", "datum_prijeti", "uzavreno", "odpovedna_osoba"):
        if key in patch:
            item[key] = patch[key] or (None if key == "uzavreno" else "")
    save_all(data)
    item["celkovy_stav"] = overall_status(item)
    item["celkem_hodin"] = minutes_to_hours(reklamace_total_minutes(item))
    return item


def update_faze(cislo: str, faze_key: str, patch: dict) -> dict | None:
    if faze_key not in FAZE_KEYS:
        return None
    data = load_all()
    item = data.get("items", {}).get(cislo)
    if not item:
        return None
    faze = item["faze"][faze_key]
    for key in ("stav", "datum", "poznamka", "odpovedna_osoba", "dodavatel",
                "dopravce", "datum_uhrady", "poznamky_faktura"):
        if key in patch:
            faze[key] = patch[key]
    for key in ("naklady_fortool", "vydana_faktura"):
        if key in patch:
            try:
                faze[key] = float(patch[key]) if patch[key] not in (None, "") else None
            except (TypeError, ValueError):
                faze[key] = None
    if "uhrada_potvrzena" in patch:
        faze["uhrada_potvrzena"] = bool(patch["uhrada_potvrzena"])
    _sync_uzavreno(item)
    save_all(data)
    item["celkovy_stav"] = overall_status(item)
    item["celkem_hodin"] = minutes_to_hours(reklamace_total_minutes(item))
    return item


def update_krok(cislo: str, faze_key: str, idx: int, patch: dict) -> dict | None:
    if faze_key not in FAZE_KEYS:
        return None
    data = load_all()
    item = data.get("items", {}).get(cislo)
    if not item:
        return None
    kroky = item["faze"][faze_key].setdefault("kroky", _empty_kroky())
    if idx < 0 or idx >= len(kroky):
        return None
    krok = kroky[idx]
    if "minuty" in patch:
        try:
            krok["minuty"] = max(0, float(patch["minuty"]))
        except (TypeError, ValueError):
            krok["minuty"] = 0
    if "nazev" in patch and krok.get("editable_nazev"):
        krok["nazev"] = (patch["nazev"] or "").strip()
    save_all(data)
    item["celkovy_stav"] = overall_status(item)
    item["celkem_hodin"] = minutes_to_hours(reklamace_total_minutes(item))
    return item


def add_krok(cislo: str, faze_key: str, nazev: str) -> dict | None:
    if faze_key not in FAZE_KEYS:
        return None
    data = load_all()
    item = data.get("items", {}).get(cislo)
    if not item:
        return None

    is_ostatni = not nazev or nazev == OSTATNI_KATALOG_LABEL
    katalog = KROK_KATALOG_BY_NAZEV.get(nazev)
    krok = {
        "nazev": "" if is_ostatni else nazev,
        "minuty": katalog["minuty"] if katalog else EXTRA_KROK_MINUTY_DEFAULT,
        "editable_nazev": is_ostatni,
    }
    item["faze"][faze_key].setdefault("kroky", _empty_kroky()).append(krok)
    save_all(data)
    item["celkovy_stav"] = overall_status(item)
    item["celkem_hodin"] = minutes_to_hours(reklamace_total_minutes(item))
    return item


def remove_krok(cislo: str, faze_key: str, idx: int) -> dict | None:
    if faze_key not in FAZE_KEYS:
        return None
    data = load_all()
    item = data.get("items", {}).get(cislo)
    if not item:
        return None
    kroky = item["faze"][faze_key].setdefault("kroky", _empty_kroky())
    if idx < 0 or idx >= len(kroky):
        return None
    kroky.pop(idx)
    save_all(data)
    item["celkovy_stav"] = overall_status(item)
    item["celkem_hodin"] = minutes_to_hours(reklamace_total_minutes(item))
    return item


def _priloha_dir(cislo: str, faze_key: str) -> Path:
    return UPLOAD_DIR / secure_filename(cislo) / faze_key


def priloha_path(cislo: str, faze_key: str, filename: str) -> Path:
    return _priloha_dir(cislo, faze_key) / secure_filename(filename)


def add_priloha(cislo: str, faze_key: str, file_storage) -> dict | None:
    if faze_key not in FAZE_KEYS:
        return None
    data = load_all()
    item = data.get("items", {}).get(cislo)
    if not item:
        return None

    filename = secure_filename(file_storage.filename)
    if not filename:
        return None
    target_dir = _priloha_dir(cislo, faze_key)
    target_dir.mkdir(parents=True, exist_ok=True)
    file_storage.save(target_dir / filename)

    prilohy = item["faze"][faze_key]["prilohy"]
    if filename not in prilohy:
        prilohy.append(filename)
    save_all(data)
    item["celkovy_stav"] = overall_status(item)
    item["celkem_hodin"] = minutes_to_hours(reklamace_total_minutes(item))
    return item


def history_row(item: dict) -> dict:
    start = None
    if item.get("datum_prijeti"):
        try:
            start = date.fromisoformat(item["datum_prijeti"])
        except ValueError:
            start = None
    stav = overall_status(item)
    end = date.today()
    if item.get("uzavreno"):
        try:
            end = date.fromisoformat(item["uzavreno"])
        except ValueError:
            pass
    total_minutes = reklamace_total_minutes(item)
    dopravci = ", ".join(dict.fromkeys(
        item["faze"][k]["dopravce"] for k in FAZE_KEYS if item["faze"][k].get("dopravce")
    ))
    return {
        "cislo": item["cislo"],
        "servisni_partner": item["servisni_partner"],
        "kontaktni_osoba": item.get("kontaktni_osoba", ""),
        "odpovedna_osoba": item.get("odpovedna_osoba", ""),
        "datum_prijeti": item.get("datum_prijeti"),
        "uzavreno": item.get("uzavreno"),
        "stav": stav,
        "stav_label": STAV_LABELS_HISTORIE.get(stav, stav),
        "doba_pracovnich_dni": business_days_between(start, end) if start else None,
        "aktualni_faze": aktualni_faze_label(item),
        "celkem_minut": total_minutes,
        "celkem_hodin": minutes_to_hours(total_minutes),
        "naklady_celkem": round(reklamace_naklady_celkem(item), 2),
        "dopravci": dopravci,
        "faktura_celkem": round(reklamace_faktura_celkem(item), 2),
        "faktura_neuhrazeno": round(reklamace_faktura_neuhrazeno(item), 2),
        "uhrada_stav": reklamace_uhrada_stav(item),
        "poznamky": reklamace_poznamky_souhrn(item),
    }


def history_all(brand: str | None = None) -> list:
    return [history_row(i) for i in list_all(brand)]


def export_xlsx(brand: str, out_path: Path, cisla: list | None = None) -> Path:
    """Export VAT do Excelu — podklad pro fakturaci. Jeden řádek na
    UZAVŘENOU reklamaci se souhrnnými hodinami za každého dodavatele
    (bez rozpisu jednotlivých kroků), a souhrnné řádky Průměr/TOTAL/
    PŘEVOD hodin do dalšího měsíce/SAZBA/K FAKTURACI bez DPH. Pokud je
    zadán výběr cisla, exportují se jen ty (a stále jen uzavřené)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    items = [i for i in list_all(brand) if i["celkovy_stav"] == "vyřízeno"]
    if cisla is not None:
        vybrana = set(cisla)
        items = [i for i in items if i["cislo"] in vybrana]
    items.sort(key=lambda i: i.get("datum_prijeti") or "")
    nastaveni = get_nastaveni(brand)

    HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
    HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    THIN = Side(style="thin", color="B7B7B7")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BOLD = Font(name="Arial", size=10, bold=True)
    NORMAL = Font(name="Arial", size=10)
    CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = date.today().strftime("%m.%Y")
    ws.sheet_view.showGridLines = False

    headers = ["Číslo", "Servisní partner", "Kontaktní osoba", "Datum přijetí",
               "Datum ukončení", "Stav", "Doba trvání (prac. dny)"]
    headers += [label for _key, label in EXPORT_FAZE_SLOUPCE]
    headers.append("TOTAL (hod.)")
    total_col_idx = len(headers)
    total_col = get_column_letter(total_col_idx)

    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = CENTER_WRAP
        c.border = BORDER

    widths = [12, 28, 20, 13, 13, 12, 13] + [15] * len(EXPORT_FAZE_SLOUPCE) + [12]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    row = first_data_row = 2
    for item in items:
        datum_prijeti = item.get("datum_prijeti")
        uzavreno = item.get("uzavreno")
        start = date.fromisoformat(datum_prijeti) if datum_prijeti else None
        end = date.fromisoformat(uzavreno) if uzavreno else date.today()
        doba = business_days_between(start, end) if start else None

        vals = [item["cislo"], item["servisni_partner"], item.get("kontaktni_osoba", ""),
                start, date.fromisoformat(uzavreno) if uzavreno else None,
                STAV_LABELS_HISTORIE.get(item["celkovy_stav"], item["celkovy_stav"]), doba]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = NORMAL
            c.border = BORDER
            if ci in (4, 5):
                c.number_format = "DD.MM.YYYY"

        faze_ci = 8
        for faze_key, _label in EXPORT_FAZE_SLOUPCE:
            hodiny = minutes_to_hours(faze_total_minutes(item["faze"][faze_key]))
            c = ws.cell(row=row, column=faze_ci, value=hodiny)
            c.font = NORMAL
            c.number_format = "0.0#"
            c.border = BORDER
            faze_ci += 1

        celkem_hodin = minutes_to_hours(reklamace_total_minutes(item))
        tcell = ws.cell(row=row, column=total_col_idx, value=celkem_hodin)
        tcell.font = BOLD
        tcell.number_format = "0.0#"
        tcell.border = BORDER
        row += 1

    last_data_row = row - 1

    row += 1
    ws.cell(row=row, column=1, value="Průměr prac. dnů / uzavřenou reklamaci").font = BOLD
    if last_data_row >= first_data_row:
        avg_cell = ws.cell(row=row, column=7, value=f"=AVERAGE(G{first_data_row}:G{last_data_row})")
        avg_cell.font = BOLD
        avg_cell.number_format = "0.0#"
    row += 1
    total_row = row
    ws.cell(row=row, column=1, value="TOTAL").font = BOLD
    if last_data_row >= first_data_row:
        sum_cell = ws.cell(row=row, column=total_col_idx,
                            value=f"=SUM({total_col}{first_data_row}:{total_col}{last_data_row})")
        sum_cell.font = BOLD
        sum_cell.number_format = "0.0#"
    row += 1
    prevod_row = row
    ws.cell(row=row, column=1, value="PŘEVOD hodin do dalšího měsíce (počáteční rozpočet)").font = NORMAL
    pc = ws.cell(row=row, column=total_col_idx, value=nastaveni["prevod_hodin"])
    pc.number_format = "0.0#"
    row += 1
    ws.cell(row=row, column=1, value="ZŮSTATEK (po odečtení této fakturace)").font = BOLD
    if last_data_row >= first_data_row:
        zc = ws.cell(row=row, column=total_col_idx, value=f"={total_col}{prevod_row}-{total_col}{total_row}")
        zc.font = BOLD
        zc.number_format = "0.0#"
    row += 1
    sazba_row = row
    ws.cell(row=row, column=1, value="SAZBA").font = NORMAL
    sc = ws.cell(row=row, column=total_col_idx, value=None)
    sc.number_format = "#,##0"
    row += 1
    ws.cell(row=row, column=1, value="K FAKTURACI bez DPH").font = BOLD
    fakt_cell = ws.cell(row=row, column=total_col_idx, value=f"={total_col}{total_row}*{total_col}{sazba_row}")
    fakt_cell.font = BOLD
    fakt_cell.number_format = "#,##0"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def remove_priloha(cislo: str, faze_key: str, filename: str) -> dict | None:
    if faze_key not in FAZE_KEYS:
        return None
    data = load_all()
    item = data.get("items", {}).get(cislo)
    if not item:
        return None

    filename = secure_filename(filename)
    prilohy = item["faze"][faze_key]["prilohy"]
    if filename in prilohy:
        prilohy.remove(filename)
        path = priloha_path(cislo, faze_key, filename)
        if path.exists():
            path.unlink()
    save_all(data)
    item["celkovy_stav"] = overall_status(item)
    item["celkem_hodin"] = minutes_to_hours(reklamace_total_minutes(item))
    return item
